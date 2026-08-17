import sys
import openpyxl
import json

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook(r'C:\Users\maxiv\.gemini\antigravity-ide\scratch\ancu\padron_maestro.xlsx', data_only=True)
print("Sheet Names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} (max row: {sheet.max_row}, max col: {sheet.max_column}) ---")
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("Empty sheet.")
        continue
    
    # Print first 10 rows
    for idx, row in enumerate(rows[:15]):
        print(f"Row {idx+1}: {row}")

wb.close()
